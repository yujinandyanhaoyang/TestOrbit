import copy
import datetime
import json
import os
import time
from urllib.parse import urlencode

import requests
from django.db.models import Max, F
from django.db.models.functions import JSONObject
from openpyxl import load_workbook
from requests import ReadTimeout
from rest_framework.response import Response

from apiData.models import ApiCaseStep, ApiCase, ApiForeachStep
from utils.comDef import get_proj_envir_db_data, db_connect, execute_sql_func, \
    close_db_con, json_dumps, JSONEncoder, MyThread, json_loads, format_parm_type_v
from utils.constant import USER_API, VAR_PARAM, HEADER_PARAM, HOST_PARAM, RUNNING, SUCCESS, FAILED, DISABLED, \
    INTERRUPT, SKIP, API_CASE, API_FOREACH, TABLE_MODE, STRING, DIY_CFG, JSON_MODE, PY_TO_CONF_TYPE, CODE_MODE, \
    OBJECT, FAILED_STOP, WAITING, PRO_CFG, FORM_MODE, EQUAL, API_VAR, NOT_EQUAL, \
    CONTAIN, NOT_CONTAIN, TEXT_MODE, API, FORM_FILE_TYPE, FORM_TEXT_TYPE, API_SQL, RES_BODY
from utils.diyException import DiyBaseException, NotFoundFileError
from utils.paramsDef import parse_param_value, run_params_code, parse_temp_params, get_parm_v_by_temp
from config.models import Environment
from user.models import UserCfg, UserTempParams

# 功能函数切分保存位置,变更到其他位置
from .function.steps_def import go_step


class ApiCasesActuator:
    """
    接口用例执行器
    user_id：执行计划的用户，传递了user_id时才会进行中断判断(调试时不会传递)
    temp_params:默认参数
    """

    def __init__(self, user_id, trigger_case_id=None, cfg_data=None, temp_params=None):
        self.user_id = user_id
        self.timeout = 60  # 默认接口超时时间
        self.base_params_source = {'user_id': self.user_id, 'case_id': trigger_case_id}
        user_cfg = UserCfg.objects.filter(user_id=user_id).values().first() or {
            'envir_id': 1, 'failed_stop': False, 'only_failed_log': False}
        cfg_data = {**user_cfg, **cfg_data} if cfg_data else user_cfg
        self.envir = cfg_data['envir_id']
        self.failed_stop = cfg_data['failed_stop']
        self.only_failed_log = cfg_data['only_failed_log']
        self.status = SUCCESS
        self.cascader_error = False
        temp_params = temp_params or UserTempParams.objects.filter(user_id=user_id).values()
        params = parse_temp_params(temp_params)
        self.default_header, self.default_var, self.default_host = (params[key] for key in ('header', 'var', 'host'))
        self.params_source = params['params_source']
        self.user_id = user_id  # 执行计划的用户，传递了user_id时才会进行中断判断
        self.status = RUNNING  # 初始化执行状态为执行中
        self.api_data = {}  # 执行过的接口信息会存在这，避免频繁查库，示例:{id:{'path':/xx,'method':'GET','timeout':10}}
        self.api_process = ''

    @staticmethod
    def clear_upload_files(upload_files_list):
        """
        清除请求中上传的临时文件
        """
        for file in upload_files_list:
            file['file'] and file['file'].close()
            os.remove(file['name'])

    def parse_api_step_output(self, params, prefix_label, step_name, response, res_headers, i):
        """
        处理api步骤的输出
        """
        out_data = {}
        if output := params.get('output_source'):
            self.api_process = '【输出参数】'
            if params['output_mode'] == TABLE_MODE:
                for out in output:
                    out_v, out_name = out['value'], str(parse_param_value(out['name'], self.default_var, i)).strip('?')
                    is_assert = not out['name'].endswith('?')
                    # 兼容老版本2023年8月15日
                    if isinstance(out_v, str):
                        locate_v = out_v
                        res_source = response
                    else:
                        # 确保从字典中安全获取value
                        if isinstance(out_v, dict) and 'value' in out_v:
                            locate_v = out_v['value']
                            res_source = response if out_v.get('source') == RES_BODY else res_headers
                        else:
                            locate_v = out_v
                            res_source = response
                    
                    # 处理字符串和其他类型情况
                    if isinstance(locate_v, str) and '.' in locate_v:
                        # 字符串且包含点号，按点分割
                        value_location_list = [parse_param_value(var, self.default_var, i) for var in locate_v.split('.')]
                    else:
                        # 如果locate_v是其他类型或不包含点的字符串，直接作为整体使用
                        value_location_list = [parse_param_value(locate_v, self.default_var, i)]
                    try:
                        res = get_parm_v_by_temp(value_location_list, res_source)
                    except Exception as e:
                        if is_assert:
                            return {'status': FAILED, 'results': str(e)}
                        else:
                            res = False
                    if not res:
                        if is_assert:
                            return {'status': FAILED, 'results': locate_v + ':' + '未在响应中找到！'}
                    else:
                        res_v = res['value']
                        self.default_var[out_name] = res_v
                        out_data[out_name] = res_v
                        self.params_source[VAR_PARAM][out_name] = {
                            'name': out_name, 'value': res_v, 'step_name': prefix_label + step_name,
                            'type': VAR_PARAM, 'param_type_id': PY_TO_CONF_TYPE.get(str(type(res_v)), STRING),
                            **self.base_params_source}
            else:  # 代码模式
                res = run_params_code(output, self.default_var, i, response, res_headers)
                if isinstance(res, dict):
                    self.default_var.update(res)
                    out_data = res
                    for name, v in res.items():
                        self.params_source[VAR_PARAM][name] = {
                            'name': name, 'value': v, 'step_name': prefix_label + step_name,
                            'type': VAR_PARAM,
                            'param_type_id': PY_TO_CONF_TYPE.get(str(type(v)), STRING),
                            **self.base_params_source}
                else:
                    return {'status': FAILED, 'results': '返回数据格式不符合要求！'}
        return {'status': SUCCESS, 'out_data': out_data}

    def parse_api_step_expect(self, params, response, res_headers, i):
        """
        处理api步骤的期望
        """
        old_default_var = copy.deepcopy(self.default_var)
        if expect := params.get('expect_source'):
            self.api_process = '【预期结果】'
            if params['expect_mode'] == TABLE_MODE:
                for ext in expect:
                    ext_name, ext_v = ext['name'], parse_param_value(ext['value'], old_default_var, i)
                    ext_v_type = ext.get('type', {}).get('type') or STRING
                    ext_v = format_parm_type_v(ext_v, ext_v_type)
                    rule = ext.get('rule', EQUAL)
                    # 使用str(parse_param_value(var, var_dict))是确保键为数字时能够转换为字符串数字进行匹配
                    ext_name_list = [str(parse_param_value(name, old_default_var, i)) for name in ext_name.split('.')]
                    res = get_parm_v_by_temp(ext_name_list, response)
                    if res:
                        res_v = res['value']
                        # res_v = str(res['value']) if ext_v_type == STRING else json_loads(res['value'])
                        if rule == EQUAL and res_v != ext_v:
                            return {'status': FAILED, 'results': ext_name + '的期望值:' + str(
                                ext_v) + ' 与响应结果不相等！响应结果值为:' + str(res_v) + ';\n'}
                        elif rule == NOT_EQUAL and res_v == ext_v:
                            return {'status': FAILED, 'results': ext_name + '的期望值:' + str(
                                ext_v) + ' 与响应结果相等【期望不相等】！响应结果值为:' + str(res_v) + ';\n'}
                        elif rule == CONTAIN and ext_v not in res_v:
                            return {'status': FAILED, 'results': ext_name + '的期望值:' + str(
                                ext_v) + ' 响应结果不包含该期望！响应结果值为:' + str(res_v) + ';\n'}
                        elif rule == NOT_CONTAIN and ext_v in res_v:
                            return {'status': FAILED, 'results': ext_name + '的期望值:' + str(
                                ext_v) + ' 响应结果包含该期望【期望不包含】！响应结果值为:' + str(res_v) + ';\n'}
                    else:
                        return {'status': FAILED, 'results': '未在响应中找到字段：' + ext_name + '\n'}
            else:  # 代码模式
                res = run_params_code(expect, old_default_var, i, response, res_headers)
                if res is not None:
                    if isinstance(res, tuple) and res[0] is False:
                        return {'status': FAILED, 'results': res[1]}
                    elif res is False:
                        return {'status': FAILED, 'results': '不符合预期！'}
        return {'status': SUCCESS}

    def parse_excel_var_params(self, file_name):
        """
        解析form-data中excel存在的变量
        """
        workbook = load_workbook(file_name)
        sheets = workbook.sheetnames
        for sheet in sheets:
            ws = workbook[sheet]
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row, col)
                    cell.value = parse_param_value(cell.value, self.default_var)
        workbook.save(file_name)

    def parse_form_data_params(self, body, files_list):
        """
        解析form-data参数
        """
        req_data, body_log = {}, {}
        for key in body:
            field_data = body[key]
            if isinstance(field_data, dict) and 'type' in field_data:
                if field_data['type'] == FORM_FILE_TYPE:
                    file_name, file_url = field_data['name'], field_data['value']
                    r = requests.get(file_url)
                    if r.status_code == 404:
                        raise NotFoundFileError('未找到上传的文件：' + file_name)
                    with open(file_name, 'wb') as f:
                        f.write(r.content)
                    file_type = os.path.splitext(file_name)[-1]
                    if file_type in ['.xlsx'] and '有变量' in file_name:
                        try:
                            self.parse_excel_var_params(file_name)
                        except Exception as e:
                            files_list.append({'name': file_name, 'file': None})
                            raise Exception('导入报错：' + str(e))
                    file = open(file_name, 'rb')
                    req_data[key] = (file_name, file)
                    body_log[key] = file_name
                    files_list.append({'name': file_name, 'file': file})
                else:
                    body_log[key] = field_data['value']
                    req_data[key] = (None, field_data['value'])
            else:
                body_log = '不是有效的form-data参数！'
                break
        return req_data, body_log

    def api(self, step, prefix_label, i=0):
        """
        执行类型为接口的步骤
        优化后：参数通过关联的ApiData获取，不再使用step['params']
        """
        print("\n" + "="*60)
        print("🌐 API方法开始执行")
        print(f"📋 步骤名称: {step.get('step_name', '未命名')}")
        
        # 临时文件列表
        upload_files_list = []
        print("🗂️ 初始化上传文件列表")
    
        
        if step.get('quote_step_id') and isinstance(step.get('quote_step_id'), int):
            # 引用的步骤，通过quote_step_id获取参数，而不是step['params']
            quote_step_id = step['quote_step_id']
            print(f"📌 步骤关联了API数据，ID: {quote_step_id}")
            
            # 检查缓存
            api_base = self.api_data.get(quote_step_id)
            if api_base:
                print("✅ 从缓存获取API基础数据")
            else:
                print("🔄 缓存未命中，从数据库查询API数据...")
                # 从数据库获取API数据
                api_instance = ApiCaseStep.objects.filter(id=quote_step_id).select_related('env').first()
                
                if not api_instance:
                    print(f"❌ 数据库中未找到API数据(ID: {quote_step_id})")
                    return {'status': FAILED, 'data': f'找不到API数据(ID: {quote_step_id})'}
                
                print("✅ 数据库查询成功")
                # 构建api_base数据
                api_base = {
                    'path': api_instance.path,
                    'method': api_instance.method,
                    'timeout': api_instance.timeout or self.timeout,
                    'env_url': api_instance.env.url if api_instance.env else ''
                }
                
                # 缓存API基础数据
                self.api_data[quote_step_id] = api_base
                print("📦 API基础数据已缓存")
            
            # 获取参数
            print("🔍 获取API参数...")
            params = api_instance.params or {} if 'api_instance' in locals() else {}
            print(f"📦 参数大小: 约 {len(str(params))} 字符")
            
            # 从api_base获取基础API信息
            url_path = api_base['path']
            method = api_base['method']
            timeout = api_base.get('timeout', self.timeout)
            
            # 确定主机地址
            if api_base.get('env_url'):
                host = api_base['env_url']
            elif host := params.get('host') or '':
                if params.get('host_type') == PRO_CFG:
                    environment = Environment.objects.filter(id=self.envir).first()
                    host = environment.url if environment else ''
            elif self.default_host:
                host = self.default_host
            else:
                host = ''
                
        else:
            # 不存在引用步骤，可以直接使用step['params']
            print("步骤未关联API数据，使用step['params']")
            params = step.get('params', {})
            if host := params.get('host') or '':
                if params.get('host_type') == PRO_CFG:
                    environment = Environment.objects.filter(id=self.envir).first()
                    host = environment.url if environment else ''
            elif self.default_host:
                host = self.default_host
            else:
                host = ''
                
            url_path = params.get('path', '')
            method = params.get('method', 'GET')
            timeout = params.get('timeout', self.timeout)
        
        
        # print('开始拼接URL...')
        url = host + url_path
        req_log = {'url': url, 'method': method, 'response': '无响应结果', 'res_header': '无响应头'}
        res_status, results = FAILED, ''
        try:
            print('开始封装请求数据...')
            self.api_process = '【Header(请求头)】'
            if header_source := params.get('header_source'):
                header = self.parse_source_params(header_source, params.get('header_mode', 'raw'), i)
                header = {str(key).lower(): str(header[key]) for key in header}  # header的key全部转换为小写
                if not header.get('content-type'):
                    header['content-type'] = 'application/json'
                # 只有没有默认请求头时才将自定义的请求头设置为默认请求头，如果使用了全局参数且有默认请求头，则永远不会替换
                if not self.default_header:
                    self.default_header = copy.deepcopy(header)
            else:
                header = copy.deepcopy(self.default_header) or {'content-type': 'application/json'}
            self.api_process = '【query(url参数)】'
            # 处理query参数
            query = self.parse_source_params(params.get('query_source'), params.get('query_mode', 'raw'), i)
            self.api_process = '【Body(请求体)】'
            # 处理body参数
            if params.get('body_mode', 'raw') != FORM_MODE:
                body = self.parse_source_params(params.get('body_source'), params.get('body_mode', 'raw'), i)
            else:
                body = self.parse_source_params(
                    params.get('body_source'), params.get('body_mode', 'raw'), i, file_list=upload_files_list)

            # 封装request请求的请求参数    
            req_params = {'url': url, 'headers': header, 'params': query, 'method': method.lower(),
                          'allow_redirects': not params.get('ban_redirects', False), 'timeout': timeout}
            req_log.update({'header': copy.deepcopy(header), 'body': body})
            content_type = header['content-type']
            if params.get('body_mode', 'raw') != FORM_MODE:
                if 'application/json' in content_type:
                    req_params['data'] = json_dumps(body).encode() if not isinstance(body, str) else body.encode(
                        'utf-8')
                elif 'text/html' in content_type:
                    req_params['data'] = body.encode('utf-8') if isinstance(body, str) else ''
                elif 'urlencoded' in content_type or 'text/plain' in content_type:
                    if not isinstance(body, dict):
                        req_params['data'] = body
                    else:
                        req_data = {k: json.dumps(body[k], ensure_ascii=False, separators=(',', ':')) if isinstance(
                            body[k], dict) else body[k] for k in body}
                        urlencode_v = urlencode(req_data).replace('+', '%20')
                        req_params['data'] = urlencode_v
            else:
                header.pop('content-type', None)
                req_log['header']['content-type'] = 'multipart/form-data'
                req_params['files'], req_log['body'] = body
            print('http请求参数封装完毕')
            try:
                # 发送请求
                print("🚀 实际发送HTTP请求...")
                r = requests.request(**req_params)

            except KeyError as e:
                req_log['results'] = results = self.api_process + '未找到key：' + str(e)
            except (requests.exceptions.ConnectionError, ReadTimeout):
                req_log['response'] = results = '请求超时！'
            except requests.exceptions.InvalidSchema:
                req_log['results'] = results = '无效的请求地址！'
            except requests.exceptions.MissingSchema:
                req_log['results'] = results = '请求地址不能为空！'
            else:
                spend_time = float('%.2f' % r.elapsed.total_seconds())
                res_code = r.status_code
                response = ''
                if str(res_code).startswith('2'):  # 代表请求成功
                    try:
                        response = r.json()
                    except Exception:
                        response = r.text
                    out_res = self.parse_api_step_output(
                        params, prefix_label, step.get('step_name', '未命名步骤'), response, dict(r.headers), i)
                    res_status, results = out_res['status'], out_res.get('results')
                    if res_status == FAILED:
                        results = self.api_process + results
                    elif out_data := out_res.get('out_data'):
                        req_log['output'] = out_data
                    ext_res = self.parse_api_step_expect(params, response, dict(r.headers), i)
                    if res_status != FAILED:
                        res_status = ext_res['status']
                    if ext_res['status'] == FAILED:
                        results = self.api_process + ext_res.get('results', '')
                elif res_code == 404:
                    results = '请求路径不存在！'
                else:
                    results = '请求异常！'
                # 更新请求日志
                req_log.update({
                    'url': r.url, 
                    'res_header': dict(r.headers), 
                    'response': response,
                    'spend_time': spend_time, 
                    'results': results
                })
                
                print('保存结果到ApiCaseStep')
                print()
                print(f"\n⏱️ 请求耗时: {spend_time}秒")
                print(f"🔢 状态码: {res_code}")
                print(f"📊 结果状态: {res_status}")
                if results:
                    print(f"📝 结果消息: {results}")
                
        except Exception as e:
            print(f"\n❌ API执行出错: {str(e)}")
            print(f"❌ 错误行号: {e.__traceback__.tb_lineno}")
            req_log['results'] = results = self.api_process + str(e)
            
        # 清理临时文件
        # print("\n🧹 清理临时上传文件...")
        self.clear_upload_files(upload_files_list)
        
        # 准备返回结果
        result = {'status': res_status, 'data': {'msg': results, 'request_log': req_log}}
        print("\n✅ API执行完成")
        print(f"📊 最终状态: {res_status}")
        print("="*60 + "\n")
        return result

    def case(self, step, prefix_label='', cascader_level=1, i=0):
        """
        执行类型为用例
        """

        if cascader_level > 10:  # 引用计划嵌套超过10层判断为死循环
            self.cascader_error = True
            return {'status': FAILED}
        params = step.get('params')
        step_data = parse_api_case_steps([params['case_related'][-1]], is_step=True)
        prefix_label += step['step_name'] + '-'
        res_status, step_data = run_step_groups(self, step_data, prefix_label, cascader_level, i)
        if cascader_level == 1 and self.cascader_error:
            self.cascader_error = False
            return {'status': FAILED, 'results': '步骤死循环或主计划步骤嵌套的子用例超过10层！'}
        return {'status': res_status, 'results': step_data}


    def sql(self, step, prefix_label='', i=0):
        """
          执行类型为SQL的步骤
        """
        params = step.get('params') or step
        db_data = get_proj_envir_db_data(params['sql_proj_related'], envir=self.envir)
        sql = params['sql']
        if db_data:
            if 'database' in params:
                db_data['db_database'] = params['database']
            res = db_connect(db_data)
            if res['status'] == SUCCESS:
                try:
                    sql = parse_param_value(sql, self.default_var, i)
                except DiyBaseException as e:
                    return {'status': FAILED, 'results': f'执行出错：{e}', 'sql': sql}
                sql_res = execute_sql_func(res['db_con'], sql, db_data['db_type'])
                close_db_con(res)
                if sql_res['status'] == SUCCESS:
                    if sql_var := params.get('sql_var'):
                        sql_data = sql_res['data']['sql_data']
                        self.default_var[sql_var] = sql_data
                        self.params_source[VAR_PARAM][sql_var] = {
                            'name': sql_var, 'value': json_dumps(sql_data, JSONEncoder), 'type': VAR_PARAM,
                            'step_name': prefix_label + step['step_name'], 'param_type_id': OBJECT,
                            **self.base_params_source}
                    return sql_res
                return sql_res
            return res
        return {'status': FAILED, 'results': '无效的连接！'}

    def foreach(self, step, prefix_label='', cascader_level=1, i=0):
        """
        循环控制器
        """
        if cascader_level > 15:  # 嵌套超过15层则判断为死循环
            self.cascader_error = True
            return {'status': FAILED}
        params = step['params']
        times_value, break_code = params['times'], params.get('break_code')
        if times_value.isdigit():
            for_times = int(times_value) if int(times_value) <= 999 else 999
        else:  # foreach_times可能为可迭代对象（列表）的情况
            for_times = parse_param_value(times_value, self.default_var)
            if isinstance(for_times, list):
                times = len(for_times)
                for_times = times if times <= 100000 else 100000  # 最多循环10万次
            elif for_times not in ('true', True) and not isinstance(for_times, int):
                raise DiyBaseException(f'无效的循环次数值：{for_times}！')
        if 'steps' not in params:
            steps = set_foreach_tree(ApiForeachStep.objects.filter(step_id=step['id']).values().order_by('id'))
        else:  # 调试时，不会传递foreach_id
            steps = params['steps']
        prefix_label += step['step_name'] + '-'
        res_status, res_data = SUCCESS, []
        if type(for_times) != int and for_times in ('true', True):
            while True:
                # 满足break条件的话则中止循环
                if self.status == INTERRUPT or break_code and run_params_code(
                        break_code, copy.deepcopy(self.default_var), i):
                    break
                run_status, step_data = run_step_groups(
                    self, copy.deepcopy(steps), prefix_label, cascader_level=cascader_level, i=i)
                i += 1
                res_data.append(step_data)
                if run_status == FAILED:
                    res_status = FAILED
        else:
            for _ in range(for_times):
                # 满足break条件的话则中止循环
                if self.status == INTERRUPT or break_code and run_params_code(
                        break_code, copy.deepcopy(self.default_var), i):
                    break
                run_status, step_data = run_step_groups(
                    self, copy.deepcopy(steps), prefix_label, cascader_level=cascader_level, i=i)
                i += 1
                res_data.append(step_data)
                if run_status == FAILED:
                    res_status = FAILED
        if cascader_level == 1:
            if self.cascader_error:
                self.cascader_error = False
                return {'status': FAILED, 'results': '步骤死循环或主计划步骤嵌套的子用例超过15层！'}
        return {'status': res_status, 'results': res_data}

    def parse_source_params(self, data, mode=TABLE_MODE, i=0, params_type='', file_list=None):
        """
        解析请求数据
        """
        res = {}
        if mode == TABLE_MODE:
            for param in data:
                if p_name := param.get('name'):
                    p_v = param.get('value')
                    parm_type = param.get('type', {}).get('type') or STRING
                    p_name = parse_param_value(p_name, self.default_var, i)
                    p_v = parse_param_value(p_v, self.default_var, i)
                    res[p_name] = format_parm_type_v(p_v, parm_type)
                    if params_type == API_VAR:
                        self.default_var[p_name] = res[p_name]
        elif mode == JSON_MODE:
            if isinstance(data, dict):
                for p_name, p_v in data.items():
                    p_name = parse_param_value(p_name, self.default_var, i)
                    res[p_name] = parse_param_value(p_v, self.default_var, i)
                    if params_type == API_VAR:
                        self.default_var[p_name] = res[p_name]
            else:
                res = parse_param_value(data, self.default_var, i)
        elif mode == TEXT_MODE:
            res = parse_param_value(data, self.default_var, i)
        elif mode == CODE_MODE:
            res = run_params_code(data, copy.deepcopy(self.default_var), i)
            if isinstance(res, dict):
                if params_type == API_VAR:  # 步骤类型为全局变量的话，则将其加入到全局变量中
                    self.default_var.update(res)
            else:
                raise DiyBaseException('返回格式不正确！需要返回一个字典')
        elif mode == FORM_MODE:
            req_data, body_log = {}, {}
            for v in data:
                parm_name, parm_v = v['name'], v.get('value') or {'value': '', 'type': FORM_TEXT_TYPE}
                if isinstance(parm_v, dict) and 'type' in parm_v:
                    if parm_v['type'] == FORM_FILE_TYPE:
                        file_name, file_url = parm_v['name'], parm_v['value']
                        r = requests.get(file_url)
                        if r.status_code == 404:
                            raise NotFoundFileError('未找到上传的文件：' + file_name)
                        with open(file_name, 'wb') as f:
                            f.write(r.content)
                        file_type = os.path.splitext(file_name)[-1]
                        if file_type in ['.xlsx'] and '有变量' in file_name:
                            try:
                                self.parse_excel_var_params(file_name)
                            except Exception as e:
                                file_list.append({'name': file_name, 'file': None})
                                raise Exception('导入报错：' + str(e))
                        file = open(file_name, 'rb')
                        req_data[parm_name] = (file_name, file)
                        body_log[parm_name] = file_name
                        file_list.append({'name': file_name, 'file': file})
                    else:
                        parm_v = parse_param_value(parm_v['value'], self.default_var)
                        body_log[parm_name] = parm_v
                        req_data[parm_name] = (None, parm_v)
                else:
                    body_log = '不是有效的form-data参数！'
                    break
            res = [req_data, body_log]
        return res


def save_results(step_data, case_data):
    """
    执行完成后，写入结果
    """

    ApiCase.objects.bulk_update(case_data, fields=('status', 'report_data', 'latest_run_time'))
    ApiCaseStep.objects.bulk_update(step_data, fields=('status', 'results', 'params'))


def run_step_groups(actuator_obj, step_data, prefix_label='', cascader_level=0, i=0):
    """
    执行步骤合集
    """
    # 默认测试是通过的
    run_status = SUCCESS
    for step in step_data:
        s_type = step['type']
        if step.get('enabled'):
            params = {'actuator_obj': actuator_obj, 'step': step, 'prefix_label': prefix_label,
                      'i': i}
            if s_type in (API_CASE, API_FOREACH):
                params['cascader_level'] = cascader_level + 1
            res = go_step(**params)
            step.update(res)
        else:
            step['status'] = DISABLED
        # step.update({'status': res['status'], 'results': res.get('results')})
        # 当测试计划状态为通过且步骤状态为失败时，就将计划状态改为失败
        if run_status != FAILED and step['status'] == FAILED:
            run_status = FAILED
    return run_status, step_data




def monitor_interrupt(user_id, actuator_obj):
    while True:
        time.sleep(3)
        # 检查执行器状态和用户配置
        exec_status = UserCfg.objects.filter(user_id=user_id).values_list('exec_status', flat=True).first()
        
        # 如果执行器已完成或用户要求中断，则停止监控
        if actuator_obj.status not in (RUNNING, WAITING) or exec_status in (INTERRUPT, WAITING):
            print('监控线程结束，状态:', actuator_obj.status)
            if exec_status == INTERRUPT:
                actuator_obj.status = INTERRUPT
            break
            
        # 只有在调试模式或需要时输出状态
        # print('monitor_interrupt', actuator_obj.status)


def run_api_case_func(case_data, user_id, cfg_data=None, temp_params=None):
    """
    执行api用例的主方法
    执行测试计划：case_data={case_id:[step1,step2,step3]}
    实时调试/步骤中计划：case_data=[step1,step2,step3]
    temp_params为空的话则查询用户的参数来测试。
    """

    res_step_objs, res_case_objs = [], []
    actuator_obj = ApiCasesActuator(user_id, cfg_data=cfg_data, temp_params=temp_params)
    thread = MyThread(target=monitor_interrupt, args=[user_id, actuator_obj])
    thread.start()
    if isinstance(case_data, dict):
        for case_id, v in case_data.items():
            print(f'这是{case_id}号用例')
            start_time = datetime.datetime.now()
            case_objs = ApiCase.objects.filter(id=case_id).first()
            if case_objs:
                case_objs.status = RUNNING
                case_objs.save(update_fields=['status'])
            report_dict = {'envir': actuator_obj.envir, 'start_time': start_time.strftime('%Y-%m-%d %H:%M:%S'),
                           'steps': []}
            actuator_obj.base_params_source['case_id'] = case_id
            # 默认测试是通过的
            case_status, step_data = run_step_groups(actuator_obj, v)
            report_dict['steps'] = step_data
            for step in step_data:
                # 过滤掉不属于ApiCaseStep模型的字段
                valid_fields = {
                    'id', 'type', 'enabled', 'step_name', 'step_order', 'status', 
                    'retried_times', 'controller_data', 'params', 'results', 
                    'timeout', 'source', 'case_id'
                }
                filtered_step = {k: v for k, v in step.items() if k in valid_fields}
                
                # 将执行结果(data字段)存储到results字段中
                if 'data' in step and step['data']:
                    filtered_step['results'] = step['data']
                
                # 确保case_id字段存在
                filtered_step['case_id'] = case_id
                res_step_objs.append(ApiCaseStep(**filtered_step))
            end_time = datetime.datetime.now()
            report_dict['spend_time'] = format((end_time - start_time).total_seconds(), '.1f')
            if actuator_obj.status in (INTERRUPT, FAILED_STOP):
                case_status = actuator_obj.status
            res_case_objs.append(
                ApiCase(id=case_id, status=case_status, latest_run_time=end_time, report_data=report_dict))
        save_results(res_step_objs, res_case_objs)
        print(f"已完成{case_id}号用例的执行")
        
        # 确保执行状态设置为WAITING，通知监控线程可以终止
        UserCfg.objects.filter(user_id=user_id).update(exec_status=WAITING)
        
        return {'params_source': actuator_obj.params_source}


def parse_api_case_steps(case_ids=None, is_step=False):
    """
    转化API计划步骤
    is_step:false代表非步骤中的用例，即外层计划列表中选中执行的用例
    """
    step_data = []
    if case_ids:
        # 参数现在通过关联的ApiData.params获取
        step_data = list(ApiCaseStep.objects.filter(case_id__in=case_ids).select_related(
            'case', 'case__module').values(
            'case_id', 'step_order', 'step_name', 'type', 'status', 'results', 'id',
            'controller_data', 'enabled','params').order_by('case_id', 'step_order'))
        
        
        if not is_step:  # 如果非测试计划步骤而是执行测试用例，需要转为{case_id:[step,step],case_id2:[step,step]}的形式
            case_data = {case_id: [] for case_id in case_ids}  # {case1:steps,case2:steps}
            for step in step_data:
                case_data[step['case_id']].append(step)
            return case_data
    return step_data


def parse_create_foreach_steps(save_step_objs, foreach_step, parent_step, next_order, parent_id=None):
    """
    格式化循环控制器步骤为创建数据
    parent_step: 父级ApiCaseStep实例
    next_order: 下一个步骤的顺序号
    """
    for step in foreach_step:
        step.pop('results', None)
        if (s_type := step['type']) == API:
            step['api_id'] = step['params']['api_id']
        elif s_type == API_CASE:
            step['quote_case_id'] = step['params']['case_related'][-1]
        step.update({'step': parent_step, 'step_order': next_order, 'parent_id': parent_id})
        save_step_objs.append(ApiForeachStep(**step))
        next_order += 1
        if s_type == API_FOREACH:
            next_order = parse_create_foreach_steps(save_step_objs, step['params'].pop('steps', []), parent_step, next_order, step['step_order'])
    return next_order


def set_foreach_tree(_list):
    """
    生成循环控制器树
    """
    print("🌳 set_foreach_tree 开始")
    print(f"🌳 输入数据: {_list}")
    print(f"🌳 输入数据长度: {len(_list) if _list else 0}")
    
    _dict, tree = {}, []
    
    # 第一遍遍历：建立字典映射并初始化foreach步骤
    print("🌳 第一遍遍历：建立映射...")
    for i in _list:
        print(f"🌳 处理项目: {i}")
        _dict[i['id']] = i
        if i['type'] == API_FOREACH:
            i['params']['steps'] = []
            print(f"🌳 初始化 foreach 步骤 {i['id']} 的 steps 数组")
    
    print(f"🌳 字典映射完成，共 {len(_dict)} 项")
    
    # 第二遍遍历：建立父子关系
    print("🌳 第二遍遍历：建立父子关系...")
    for i in _list:
        node = i
        parent_id = node['parent_id']
        print(f"🌳 节点 {node['id']} 的父节点: {parent_id}")
        
        if parent_id is not None:
            if parent_id in _dict:
                _dict[parent_id]['params']['steps'].append(node)
                print(f"🌳 ✅ 节点 {node['id']} 添加到父节点 {parent_id}")
            else:
                print(f"🌳 ❌ 父节点 {parent_id} 不存在")
        else:
            tree.append(node)
            print(f"🌳 ✅ 根节点 {node['id']} 添加到树")
    
    print(f"🌳 树构建完成，根节点数: {len(tree)}")
    print(f"🌳 最终树结构: {tree}")
    print("🌳 set_foreach_tree 完成")
    return tree

